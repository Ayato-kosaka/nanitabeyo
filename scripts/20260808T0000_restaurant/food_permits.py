#!/usr/bin/env python3
"""自治体ごとにばらばらな食品営業許可台帳を、1つの形式に揃える。

## なぜ4つ目のソースが要るのか

#1276 の PoC で、Google の店に届かない分を調べたところ、**名寄せの失敗ではなく
オープンデータ側の欠落**だった。名前が分かる未到達 45 件のうち 44 件は、最寄りの
オープンデータ行が別の店である。欠けていたのは**雑居ビル上階のスナック・バー**で、
業態プロファイルでは未到達 9.8% 対 到達 1.3%（7.6倍）と出た。

IFAS（`1_4_load_ifas.py`）は 2021年6月の食品衛生法改正**以降**にオンライン申請
された分しか持たない。改正前に許可を取った施設は各自治体が個別に公開している。
鹿児島市の実測では IFAS 9,812 行に対し、改正前の台帳が別に 4,885 行あった。

許可台帳の住所には**階数**が入る（「鹿児島市千日町15番15号4階」）。Overture と
OSM が持っていない上階テナントを指せるのはこれである。実測で ③（Google 側から
見た到達率）は台帳のある自治体で 66.5% → 78.9%、渋谷区では 84.1% になった。

## 実装上の注意

自治体ごとに列名も文字コードも公開形式も違う。**当たらなかった行は黙って捨てず、
理由つきで数える。** PoC では取り込み行数が 4,160 → 174,828 になるまでに、
数え上げを見て8回直した。以下の定数のコメントはその記録である。
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sqlite3
import sys
import unicodedata
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

ROOT = Path(__file__).resolve().parent
csv.field_size_limit(min(sys.maxsize, 2**31 - 1))

ENCODINGS = ("utf-8-sig", "cp932", "utf-8", "euc-jp")

NAME_HINTS = ("名称", "屋号", "商号", "施設名", "店名")
NAME_EXCLUDE = ("営業者", "申請者", "法人番号", "代表者", "業種", "種別", "分類")
# 施設側の住所を優先する。営業者（法人）の住所と取り違えると、空欄だらけの列を
# 住所として読んでしまう（鹿児島県のファイルで 14,123 行が落ちた）。
ADDRESS_HINTS = (
    "営業施設所在地", "営業所所在地", "施設所在地", "店舗所在地",
    "所在地", "施設住所", "営業所住所", "住所", "設置場所",
    # 宇都宮市は列名が「営業所」だけで、中身が施設の住所である。
    # 「所在地」も「住所」も付かない列名があることを見落とすと、その自治体が丸ごと落ちる。
    "営業所", "施設", "店舗",
)
# 「所在地郵便番号」を住所列と誤認して、郵便番号を住所として読んでいた
# （大分県のファイルで 18,732 行が落ちた）。番号系の列は住所ではない。
ADDRESS_EXCLUDE = (
    "郵便番号", "電話", "ｆａｘ", "fax", "コード", "番号",
    # 営業者・申請者の住所は施設の住所ではない
    "営業者", "申請者", "法人", "代表者", "届出者", "許可者",
    # 「営業所」を住所として拾うようにしたので、名称・カナ側を巻き込まないようにする
    "名称", "屋号", "商号", "カナ", "英字", "フラグ", "業種", "種別", "分類", "日", "者",
)
TYPE_HINTS = ("営業の種類", "営業の種類もしくは営業の形態", "業種", "許可業種", "種別", "業態")
CLOSED_HINTS = ("廃業", "失効")
# 「廃業日」のような列ではなく、「営業ステータス」列に「廃業」と書く形式もある。
# 渋谷区の 39,106 行は 22,122 行が廃業で、これを落とさないと半分が死んだ店になる。
STATUS_HINTS = ("営業ステータス", "ステータス", "状態", "許可状況")
CLOSED_VALUES = ("廃業", "失効", "取消", "取り消", "抹消")
LATITUDE_HINTS = ("緯度", "latitude")
LONGITUDE_HINTS = ("経度", "longitude")

# IFAS と同じ基準で飲食店に絞る。ここを変えると①③の分母が動くので合わせておく。
KEEP_KEYWORDS = ("飲食店", "喫茶店", "菓子製造", "そうざい製造", "アイスクリーム")

PREFECTURE = re.compile(r"^(東京都|北海道|(?:京都|大阪)府|.{2,3}県)")
# 台帳の住所は都道府県から始まらないことが多い（「中央区銀座1-1-1」など）。
# 実測で 18,384 行がこれで落ちていた。市区町村名から都道府県を補って救う。
MUNICIPALITY_HEAD = re.compile(r"^(.{1,6}?市.{1,5}?区|.{1,8}?[市区町村])")


def load_prefecture_map(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def with_prefecture(address: str, prefectures: dict[str, str], hint: str) -> str:
    """都道府県から始まらない住所を、公開元の自治体名で補完する。

    台帳の住所は3通りある。

    1. 「鹿児島県鹿児島市新栄町14番2号」— そのまま使える
    2. 「中央区銀座1-1-1」— 市区町村からは始まる。都道府県だけ補う
    3. 「新宿1-1-1」— 自分の区の中なので区名すら書かない。公開元の自治体名
       （ファイル名の頭）から「東京都新宿区」を丸ごと前置する

    3 を拾わないと東京23区の台帳がまるごと落ちる（実測で 17,044 行）。
    同じ市区町村名が複数県にあるとき（府中市・伊達市・東村）も、公開元の
    自治体名が優先される。
    """

    if PREFECTURE.match(address):
        return address
    source = hint.split("__")[0].strip()
    match = MUNICIPALITY_HEAD.match(address)
    if match:
        municipality = match.group(1)
        prefecture = prefectures.get(municipality)
        hinted = PREFECTURE.match(source)
        if hinted:
            prefecture = hinted.group(1)
        return f"{prefecture}{address}" if prefecture else address
    # 住所が市区町村からも始まらない。公開元が市区町村なら丸ごと前置する。
    prefecture = prefectures.get(source)
    if prefecture and address:
        return f"{prefecture}{source}{address}"
    return address


def normalise_header(value: str) -> str:
    return unicodedata.normalize("NFKC", value or "").replace(" ", "").strip()


def pick(headers: list[str], hints: tuple[str, ...], exclude: tuple[str, ...] = ()) -> str | None:
    for header in headers:
        flat = normalise_header(header)
        if any(word in flat for word in exclude):
            continue
        if any(hint in flat for hint in hints):
            return header
    return None


def read_xlsx(path: Path) -> list[dict]:
    """xlsx を1枚目のシートだけ読む。台帳の xlsx は 168 件あり、無視できない。"""

    try:
        import openpyxl
    except ImportError:
        return []
    try:
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - 自治体の xlsx は壊れていることがある
        return []
    sheet = book[book.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header: list[str] = []
    out: list[dict] = []
    for values in rows:
        cells = ["" if value is None else str(value).strip() for value in values]
        if not header:
            # 先頭に表題行が入っていることがある。列名らしい行まで読み飛ばす。
            if sum(1 for cell in cells if cell) >= 3:
                header = cells
            continue
        if not any(cells):
            continue
        out.append(dict(zip(header, cells)))
    book.close()
    return out


def read_xls(path: Path) -> list[dict]:
    """古い xls。台帳には 102 件あった。openpyxl では読めないので xlrd を使う。"""

    try:
        import xlrd
    except ImportError:
        return []
    try:
        book = xlrd.open_workbook(str(path))
    except Exception:  # noqa: BLE001
        return []
    sheet = book.sheet_by_index(0)
    header: list[str] = []
    out: list[dict] = []
    for index in range(sheet.nrows):
        cells = [str(value).strip() for value in sheet.row_values(index)]
        if not header:
            if sum(1 for cell in cells if cell) >= 3:
                header = cells
            continue
        if not any(cells):
            continue
        out.append(dict(zip(header, cells)))
    return out


def read_rows(path: Path) -> tuple[list[dict], str]:
    if path.suffix.lower() == ".xls":
        return read_xls(path), "xls"
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        rows = read_xlsx(path)
        return rows, "xlsx"
    for encoding in ENCODINGS:
        try:
            with path.open(encoding=encoding, newline="") as handle:
                rows = list(csv.DictReader(handle))
            if rows and any(rows[0].keys()):
                return rows, encoding
        except (UnicodeDecodeError, csv.Error):
            continue
    return [], ""




@dataclass(frozen=True)
class PermitRow:
    """台帳1行を、ソース非依存の形に揃えたもの。"""

    jurisdiction: str
    source_record_id: str
    name: str
    address: str
    business_type: str
    latitude: float | None
    longitude: float | None
    coordinate_source: str
    raw_payload_json: str


def iter_permit_rows(
    directory: Path,
    *,
    prefectures: dict[str, str],
    coordinates: dict[str, tuple[float, float, str]],
    stats: Counter,
) -> Iterator[PermitRow]:
    """台帳ディレクトリを読み、正規化した行を返す。

    落とした行は ``stats`` に理由つきで積む。呼び出し側は必ずこれを記録すること。
    捨てた量が見えていないと、列名の揺れ1つで自治体が丸ごと消えても気づけない。
    """

    seen: set[tuple[str, str]] = set()
    for path in sorted(directory.glob("*")):
        if path.suffix.lower() not in (".csv", ".xlsx", ".xlsm", ".xls"):
            stats["skipped_unsupported_format"] += 1
            continue
        rows, _ = read_rows(path)
        if not rows:
            stats["unreadable"] += 1
            continue
        headers = list(rows[0].keys())
        name_column = pick(headers, NAME_HINTS, NAME_EXCLUDE)
        address_column = pick(headers, ADDRESS_HINTS, ADDRESS_EXCLUDE)
        if not name_column or not address_column:
            stats["no_name_or_address_column"] += 1
            continue
        type_column = pick(headers, TYPE_HINTS)
        closed_column = pick(headers, CLOSED_HINTS)
        status_column = pick(headers, STATUS_HINTS)
        latitude_column = pick(headers, LATITUDE_HINTS)
        longitude_column = pick(headers, LONGITUDE_HINTS)
        jurisdiction = path.stem.split("__")[0]
        file_key = hashlib.sha256(path.stem.encode("utf-8")).hexdigest()[:16]
        stats["files_used"] += 1

        for index, row in enumerate(rows, start=2):
            stats["rows_read"] += 1
            if closed_column and (row.get(closed_column) or "").strip():
                stats["closed"] += 1
                continue
            if status_column and any(
                word in (row.get(status_column) or "") for word in CLOSED_VALUES
            ):
                stats["closed_by_status"] += 1
                continue
            business = (row.get(type_column) or "") if type_column else ""
            if type_column and not any(word in business for word in KEEP_KEYWORDS):
                stats["not_a_restaurant"] += 1
                continue
            name = unicodedata.normalize("NFKC", (row.get(name_column) or "").strip())
            address = unicodedata.normalize("NFKC", (row.get(address_column) or "").strip())
            address = with_prefecture(address, prefectures, path.stem)
            if not name:
                stats["no_name"] += 1
                continue
            if not PREFECTURE.match(address):
                stats["address_without_prefecture"] += 1
                continue
            key = (name, address)
            if key in seen:
                stats["duplicate"] += 1
                continue
            seen.add(key)

            latitude = longitude = None
            coordinate_source = ""
            if latitude_column and longitude_column:
                try:
                    latitude = float(row[latitude_column])
                    longitude = float(row[longitude_column])
                    coordinate_source = "source"
                except (TypeError, ValueError, KeyError):
                    latitude = longitude = None
            if latitude is None and address in coordinates:
                latitude, longitude, level = coordinates[address]
                coordinate_source = f"geocoded:{level}"
            stats["has_coordinates" if latitude is not None else "needs_geocoding"] += 1

            yield PermitRow(
                jurisdiction=jurisdiction,
                # ファイル名を48文字で切ると、同じ自治体の別ファイルどうしで
                # 衝突する（実測: 334キーが最大10ファイル分を共有し、
                # source_records と seed_catalog に重複行を生んだ）。
                # 全体のハッシュを使えば長さを抑えたまま一意になる。
                source_record_id=f"{file_key}:{index}",
                name=name,
                address=address,
                business_type=business,
                latitude=latitude,
                longitude=longitude,
                coordinate_source=coordinate_source,
                raw_payload_json=json.dumps(row, ensure_ascii=False, sort_keys=True),
            )
